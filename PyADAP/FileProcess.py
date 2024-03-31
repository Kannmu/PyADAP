"""
PyADAP
=====

PyADAP: Python Automated Data Analysis Pipeline

File Processing

Author: Kannmu
Date: 2024/3/31
License: MIT License
Repository: https://github.com/Kannmu/PyADAP

"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def ExcelFormAdj(FilePath):
    # 加载Excel文件
    wb = load_workbook(FilePath)

    for sheet_name in wb.sheetnames:
        # 获取工作表
        ws = wb[sheet_name]

        # 获取Header行
        header_row = ws[1]

        # 自动调整列宽
        for cell in header_row:
            column_letter = cell.column_letter
            column_cells = ws[column_letter]
            length = 1.2*max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_letter].width = length

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 保存修改后的Excel文件
    wb.save(FilePath)