"""
PyADAP
=====

PyADAP: Python Automated Data Analysis Pipeline

File Processing

Author: Kannmu
Date: 2024/3/31
License: MIT License
Repository: https://github.com/Kannmu/PyADAP

"""

import os
import pandas as pd
import win32ui
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import PyADAP.Data as data

def SaveDataToExcel(
    dataIns: data.Data,
    Statistics: pd.DataFrame,
    Normality: pd.DataFrame,
    Sphericity: pd.DataFrame,
    tTest: pd.DataFrame,
    OneWayANOVA: pd.DataFrame,
    TwoWayANOVA: pd.DataFrame,
    rmANOVAResults:pd.DataFrame
):
    if not dataIns.ResultsFilePath: return
    with pd.ExcelWriter(dataIns.ResultsFilePath, engine="xlsxwriter") as writer:
        # dataIns.GetDataInfo().to_excel(writer, sheet_name="Data Information", index=False)
        Statistics.to_excel(writer, sheet_name="Statistics Results", index=False)
        Normality.to_excel(writer, sheet_name="NormalTest Results", index=False)
        Sphericity.to_excel(writer, sheet_name="Sphericity Test Results", index=False)
        if(not rmANOVAResults is None): rmANOVAResults.to_excel(writer, sheet_name="RM ANOVA Results", index=False)
        # tTest.to_excel(writer, sheet_name="T-test Results", index=False)
        # OneWayANOVA.to_excel(writer, sheet_name="One-Way ANOVA Results", index=False)
        # if(not TwoWayANOVA is None): TwoWayANOVA.to_excel(writer, sheet_name="Two-Way ANOVA Results", index=False)

    ExcelPostAdj(dataIns.ResultsFilePath)


def ExcelPostAdj(FilePath):
    """
    Adjust column widths and cell alignments in an Excel file.

    Args:
        FilePath (str): The file path of the Excel file to be adjusted.

    Notes:
        - The function loads the Excel file using `openpyxl.load_workbook`.
        - For each worksheet in the workbook:
            - Get the header row (first row).
            - Automatically adjust the column width based on the maximum length of cell values in each column, with a scaling factor of 1.2.
            - Iterate through rows starting from the second row.
            - For each cell in the row, set the horizontal and vertical alignment to center.
        - After adjusting the column widths and cell alignments, save the modified Excel file.
    """
    # Load the Excel file
    wb = load_workbook(FilePath)

    for sheet_name in wb.sheetnames:
        # Get the worksheet
        ws = wb[sheet_name]

        # Get the header row
        header_row = ws[1]

        # Autofit column widths
        for cell in header_row:
            column_letter = cell.column_letter
            column_cells = ws[column_letter]
            length = 1.2 * max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_letter].width = length

        # Iterate through rows and center align cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the modified Excel file
    wb.save(FilePath)

def SelectDataFile(default_path='./'):
    """
    This function opens a file dialog for data files. The selected file path is returned by the function.
    If the user cancels the dialog or an error occurs, None is returned.

    :param default_path: The default directory to open the file dialog in.
    :return: The selected file path, or None if no file is selected or an error occurs.
    """
    try:
        # Create a file selection dialog
        dlg = win32ui.CreateFileDialog(
            1, ".xlsx;.xls;.csv", default_path, 0, "Data (*.xlsx;*.xls;*.csv)|*.xlsx;*.xls;*.csv||"
        )

        # Display the file selection dialog
        dlg.DoModal()

        # Get the selected file path
        data_path = dlg.GetPathName()

        # Validate the file path
        if not os.path.isfile(data_path):
            print(f"The selected file does not exist: {data_path}")
            return None

        # Return the selected file path
        return data_path

    except Exception as e:
        print(f"An error occurred while selecting the file: {e}")
        return None